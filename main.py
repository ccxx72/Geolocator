from geopy.geocoders import Photon
from openpyxl import load_workbook

ADDRESS_COLUMN = 5
CAP_COLUMN = 6
LOC_COLUMN = 7

LAT_COLUMN = 23
LON_COLUMN = 24
PRECISION_COLUMN = 25


def main():
    geolocator = Photon()

    wb = load_workbook(filename='File.xlsx')
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]  # selecting always the first sheet
    row_count = sheet.max_row
    column_count = sheet.max_column
    for i in range(2, row_count):
        if i % 10 == 0:
            wb.save("File.xlsx")

        address = sheet.cell(row=i, column=ADDRESS_COLUMN).value
        cap = sheet.cell(row=i, column=CAP_COLUMN).value
        loc = sheet.cell(row=i, column=LOC_COLUMN).value

        if address and cap and loc:
            full_address = f'{address} {loc} {cap} ITALIA'
            print(i, full_address)
            try:
                location = geolocator.geocode(full_address, timeout=10)
                sheet.cell(row=i, column=LAT_COLUMN).value = location.latitude
                sheet.cell(row=i, column=LON_COLUMN).value = location.longitude
                sheet.cell(row=i, column=PRECISION_COLUMN).value = ''
            except:
                try:
                    location = geolocator.geocode(f'{cap} {loc} ITALIA', timeout=10)
                    sheet.cell(row=i, column=LAT_COLUMN).value = location.latitude
                    sheet.cell(row=i, column=LON_COLUMN).value = location.longitude
                    sheet.cell(row=i, column=PRECISION_COLUMN).value = '0'
                except:
                    sheet.cell(row=i, column=LAT_COLUMN).value = ''
                    sheet.cell(row=i, column=LON_COLUMN).value = ''
                    sheet.cell(row=i, column=PRECISION_COLUMN).value = '-1'

            print(f'({sheet.cell(row=i, column=LAT_COLUMN).value}, {sheet.cell(row=i, column=LON_COLUMN).value})')

    wb.save("File.xlsx")


if __name__ == "__main__":
    main()
